#!/usr/bin/env python3
"""Convert the ACMG SF supplementary XLSX into the YAML asset shipped with the pipeline.

Run once per ACMG SF version release. The ACMG supplement itself is copyrighted
and should NOT be committed to the repo; this script lets a maintainer regenerate
the YAML from a locally-downloaded copy.

Usage:
    python build_acmg_sf_yaml.py <supplement.xlsx> --version 3.3 --output ../assets/acmg_sf_v3_3.yaml

Sources for the supplement:
    v3.2: https://pmc.ncbi.nlm.nih.gov/articles/PMC10524344/
    v3.3: https://www.gimjournal.org/article/S1098-3600(25)00101-7/fulltext

The pipeline's ClinVar filter uses the per-gene reporting_criterion to decide
which P/LP variants to surface (e.g., HFE only reports p.C282Y homozygotes,
TTN only reports truncating variants).
"""

from __future__ import annotations

import argparse
import hashlib
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

try:
    import yaml
    from openpyxl import load_workbook
except ImportError as exc:
    print(f"ERROR: missing dependency ({exc}). Install with `pip install pyyaml openpyxl`.",
          file=sys.stderr)
    sys.exit(1)


# Map the free-text "Variants to report" strings from the supplement onto
# stable identifiers the pipeline's ClinVar filter can act on.
REPORTING_CRITERIA = {
    "All P and LP": {
        "id": "all_plp",
        "description": "Report any Pathogenic or Likely-Pathogenic variant in this gene.",
    },
    "All hemi or homozygous P and LP or 2 het. P/LP variants": {
        "id": "hemi_homo_or_2het_plp",
        "description": (
            "For X-linked or recessive disease: report hemizygous, homozygous, "
            "or compound-heterozygous (two het) P/LP variants. Single hets are not actionable."
        ),
    },
    "All hemi, het, homozygous P and LP": {
        "id": "any_zygosity_plp",
        "description": "X-linked dominant or similar: any zygosity is actionable.",
    },
    "P and LP (2 variants)": {
        "id": "two_plp_variants",
        "description": "Recessive condition: report only when at least two P/LP variants are present (compound het or homozygous).",
    },
    "P and LP (truncating variants only)": {
        "id": "truncating_plp_only",
        "description": "Report only P/LP variants that result in protein truncation (frameshift, stop-gain, splice). Missense P/LP not actionable here.",
    },
    "p.C282Y homozygotes only": {
        "id": "hfe_c282y_homozygotes",
        "description": "HFE-specific: report only homozygous p.C282Y (rs1800562). Compound hets with H63D not reported.",
    },
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_xlsx(path: Path) -> tuple[list[tuple], str]:
    wb = load_workbook(path, data_only=True)
    sheet = wb.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    # Row 0 disclaimer, Row 1 blank, Row 2 header, Row 3+ data
    header = [str(c or "").strip() for c in rows[2]]
    expected = ["Gene", "Gene MIM", "Disease/Phentyope", "Disorder MIM",
                "Phenotype Category", "Inheritance", "SF List Version", "Variants to report"]
    if [h.strip() for h in header] != expected:
        print(f"WARNING: unexpected header columns: {header}", file=sys.stderr)
    # Drop trailing disclaimer rows: a real gene row always has a numeric Gene MIM in col 1.
    data = [r for r in rows[3:] if r and r[0] and to_int_or_none(r[1]) is not None]
    return data, sheet.title


def to_int_or_none(value) -> int | None:
    """The supplement uses 'n/a' (and blanks) for missing MIM numbers."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {"n/a", "na", "none", "-"}:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def build_yaml(data: list[tuple], version: str, source_path: Path, sheet_name: str) -> dict:
    by_gene = defaultdict(list)
    for row in data:
        gene, gene_mim, disease, disorder_mim, category, inheritance, sf_version, variants_to_report = row[:8]
        by_gene[gene].append({
            "gene_mim": to_int_or_none(gene_mim),
            "disease": disease,
            "disorder_mim": to_int_or_none(disorder_mim),
            "category": (category or "").replace("\n", "; ").strip(),
            "inheritance": (inheritance or "").strip(),
            "added_in_version": str(sf_version) if sf_version is not None else None,
            "variants_to_report_raw": variants_to_report,
        })

    genes_out = []
    for symbol in sorted(by_gene):
        entries = by_gene[symbol]
        # Gene-level fields are taken from the first entry; per-disease fields stay per-entry.
        first = entries[0]
        diseases = []
        for e in entries:
            raw = e["variants_to_report_raw"]
            crit = REPORTING_CRITERIA.get(raw)
            if not crit:
                print(f"WARNING: unmapped 'Variants to report' for {symbol}: {raw!r}",
                      file=sys.stderr)
                criterion_id = "unknown"
            else:
                criterion_id = crit["id"]
            diseases.append({
                "name": e["disease"],
                "disorder_mim": e["disorder_mim"],
                "category": e["category"],
                "reporting_criterion": criterion_id,
            })
        genes_out.append({
            "symbol": symbol,
            "gene_mim": first["gene_mim"],
            "inheritance": first["inheritance"],
            "added_in_version": first["added_in_version"],
            "diseases": diseases,
        })

    return {
        "source": {
            "title": f"ACMG SF v{version}",
            "supplement_file": source_path.name,
            "supplement_sha256": sha256(source_path),
            "sheet": sheet_name,
            "generated_at": str(date.today()),
            "generator": "bin/build_acmg_sf_yaml.py",
            "note": (
                "Generated from the ACMG-distributed supplementary XLSX. "
                "The supplement itself is NOT redistributed in this repo "
                "(ACMG copyright); regenerate locally if updating."
            ),
        },
        "version": version,
        "total_genes": len(genes_out),
        "total_gene_disease_pairs": len(data),
        "reporting_criteria": [
            {"id": c["id"], "label": label, "description": c["description"]}
            for label, c in REPORTING_CRITERIA.items()
        ],
        "genes": genes_out,
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("supplement", type=Path, help="Path to the ACMG SF supplement XLSX")
    p.add_argument("--version", required=True, help="ACMG SF version string, e.g. 3.3")
    p.add_argument("--output", type=Path, required=True, help="Where to write the YAML")
    args = p.parse_args()

    if not args.supplement.exists():
        print(f"ERROR: supplement not found: {args.supplement}", file=sys.stderr)
        return 1

    data, sheet = parse_xlsx(args.supplement)
    out = build_yaml(data, args.version, args.supplement, sheet)

    with open(args.output, "w") as fh:
        fh.write(
            "# Auto-generated by bin/build_acmg_sf_yaml.py — do not edit by hand.\n"
            "# Regenerate from the official ACMG supplementary XLSX after each version release.\n"
            "# See bin/build_acmg_sf_yaml.py for sources and re-generation instructions.\n\n"
        )
        yaml.safe_dump(out, fh, sort_keys=False, width=200, allow_unicode=True)
    print(f"Wrote {args.output} ({out['total_genes']} unique genes, "
          f"{out['total_gene_disease_pairs']} gene-disease pairs).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
